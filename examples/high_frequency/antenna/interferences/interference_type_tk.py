import os
import sys
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
from PIL import ImageGrab
import openpyxl
from openpyxl.styles import PatternFill

# AEDT / EMIT
import ansys.aedt.core
from ansys.aedt.core.emit_core.emit_constants import InterfererType


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        # Main window config
        MainWindow.title("Interference Classification")
        MainWindow.geometry("550x650")

        # Root container
        self._root = ttk.Frame(MainWindow)
        self._root.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Top row: path + design name
        top_row = ttk.Frame(self._root)
        top_row.pack(fill=tk.X, pady=(0, 8))

        # Path group
        self.path_box = ttk.LabelFrame(top_row, text="Path to EMIT Project")
        self.path_box.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        path_inner = ttk.Frame(self.path_box)
        path_inner.pack(fill=tk.X, padx=6, pady=6)

        self.file_path_var = tk.StringVar()
        self.file_path_box = ttk.Entry(path_inner, textvariable=self.file_path_var, state="readonly")
        self.file_path_box.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.file_select_btn = ttk.Button(path_inner, text="...")
        self.file_select_btn.pack(side=tk.LEFT, padx=(6, 0))

        # Design group
        self.design_name_box = ttk.LabelFrame(top_row, text="Design Name")
        self.design_name_box.pack(side=tk.LEFT, fill=tk.X, padx=(0, 0))
        design_inner = ttk.Frame(self.design_name_box)
        design_inner.pack(fill=tk.X, padx=6, pady=6)

        self.design_name_dropdown = ttk.Combobox(design_inner, state="disabled", values=[])
        self.design_name_dropdown.pack(fill=tk.X)

        # Warnings label
        self.warnings_var = tk.StringVar(value="Warnings")
        self.warnings = ttk.Label(self._root, textvariable=self.warnings_var, foreground="#b30000")
        self.warnings.pack(fill=tk.X)
        self.warnings.pack_forget()

        # Tabs
        self.tab_widget = ttk.Notebook(self._root)
        self.tab_widget.pack(fill=tk.BOTH, expand=True, pady=(6, 0))

        # Protection Level tab
        self.protection_tab = ttk.Frame(self.tab_widget)
        self.tab_widget.add(self.protection_tab, text="Protection Level")

        prot_top = ttk.Frame(self.protection_tab)
        prot_top.pack(fill=tk.X, padx=6, pady=6)

        # Protection Level Thresholds
        self.protection_select_box = ttk.LabelFrame(prot_top, text="Protection Level Thresholds")
        self.protection_select_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 6))
        sel_inner = ttk.Frame(self.protection_select_box)
        sel_inner.pack(fill=tk.BOTH, padx=6, pady=6)

        self.damage_var = tk.BooleanVar(value=True)
        self.overload_var = tk.BooleanVar(value=True)
        self.intermodulation_var = tk.BooleanVar(value=True)
        self.desensitization_var = tk.BooleanVar(value=True)

        self.damage_check = ttk.Checkbutton(sel_inner, text="Damage", variable=self.damage_var)
        self.overload_check = ttk.Checkbutton(sel_inner, text="Overload", variable=self.overload_var)
        self.intermodulation_check = ttk.Checkbutton(sel_inner, text="Intermodulation", variable=self.intermodulation_var)
        self.desensitization_check = ttk.Checkbutton(sel_inner, text="Desensitization", variable=self.desensitization_var)

        self.damage_check.pack(anchor=tk.W)
        self.overload_check.pack(anchor=tk.W)
        self.intermodulation_check.pack(anchor=tk.W)
        self.desensitization_check.pack(anchor=tk.W)

        # Protection Level Classification
        self.protection_class_box = ttk.LabelFrame(prot_top, text="Protection Level Classification")
        self.protection_class_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        class_inner = ttk.Frame(self.protection_class_box)
        class_inner.pack(fill=tk.BOTH, padx=6, pady=6)

        self.radio_dropdown = ttk.Combobox(class_inner, state="disabled", values=[])
        self.radio_dropdown.pack(fill=tk.X, pady=(0, 6))

        # Legend table: use 2 columns to emulate vertical headers (Label | Value)
        self.protection_legend_table = ttk.Treeview(
            class_inner,
            columns=("label", "value"),
            show="headings",
            height=4,
        )
        self.protection_legend_table.heading("label", text="")
        self.protection_legend_table.heading("value", text="Protection Level (dBm)")
        self.protection_legend_table.column("label", width=120, anchor=tk.W)
        self.protection_legend_table.column("value", width=120, anchor=tk.CENTER)
        self.protection_legend_table.pack(fill=tk.X)

        # Auto-resize legend columns to fit available width
        def _resize_prot_legend(event=None):
            total = max(self.protection_legend_table.winfo_width() - 4, 120)
            label_w = int(total * 0.45)
            value_w = total - label_w
            self.protection_legend_table.column("label", width=max(label_w, 80))
            self.protection_legend_table.column("value", width=max(value_w, 80))
        self.protection_legend_table.bind("<Configure>", _resize_prot_legend)
        _resize_prot_legend()

        # Populate legend values and approximate background colors
        legend_rows = [
            ("Damage", "30.0", "#FFA600"),
            ("Overload", "-4.0", "#FF6361"),
            ("Intermodulation", "-30.0", "#D359A2"),
            ("Desensitization", "-104.0", "#7D73CA"),
        ]
        for label, value, color in legend_rows:
            iid = self.protection_legend_table.insert("", tk.END, values=(label, value))
            self.protection_legend_table.tag_configure(label, background=color)
            self.protection_legend_table.item(iid, tags=(label,))

        # Radio-specific levels toggle
        self.radio_specific_var = tk.BooleanVar(value=False)
        self.radio_specific_levels = ttk.Checkbutton(
            class_inner,
            text="Use radio specific protection levels",
            variable=self.radio_specific_var,
        )
        self.radio_specific_levels.pack(anchor=tk.W, pady=(6, 0))

        # Results matrix container (holds tksheet or treeview)
        self.protection_matrix = ttk.Frame(self.protection_tab)
        self.protection_matrix.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))

        # Protection buttons
        prot_btns = ttk.Frame(self.protection_tab)
        prot_btns.pack(fill=tk.X, padx=6, pady=(0, 6))
        self.protection_results_btn = ttk.Button(prot_btns, text="Generate Results")
        self.protection_export_btn = ttk.Button(prot_btns, text="Export to Excel")
        self.protection_results_btn.pack(side=tk.LEFT)
        self.protection_export_btn.pack(side=tk.LEFT, padx=6)

        # Interference Type tab
        self.interference_tab = ttk.Frame(self.tab_widget)
        self.tab_widget.add(self.interference_tab, text="Interference Type")

        int_top = ttk.Frame(self.interference_tab)
        int_top.pack(fill=tk.X, padx=6, pady=6)

        # Interference Type (Source / Victim)
        self.interference_select_box = ttk.LabelFrame(int_top, text="Interference Type (Source / Victim)")
        self.interference_select_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 6))
        int_sel_inner = ttk.Frame(self.interference_select_box)
        int_sel_inner.pack(fill=tk.BOTH, padx=6, pady=6)

        self.in_in_var = tk.BooleanVar(value=True)
        self.out_in_var = tk.BooleanVar(value=True)
        self.in_out_var = tk.BooleanVar(value=True)
        self.out_out_var = tk.BooleanVar(value=True)

        self.in_in_check = ttk.Checkbutton(int_sel_inner, text="Inband / Inband", variable=self.in_in_var)
        self.out_in_check = ttk.Checkbutton(int_sel_inner, text="Out of Band / Inband", variable=self.out_in_var)
        self.in_out_check = ttk.Checkbutton(int_sel_inner, text="Inband / Out of Band", variable=self.in_out_var)
        self.out_out_check = ttk.Checkbutton(int_sel_inner, text="Out of Band / Out of Band", variable=self.out_out_var)

        self.in_in_check.pack(anchor=tk.W)
        self.out_in_check.pack(anchor=tk.W)
        self.in_out_check.pack(anchor=tk.W)
        self.out_out_check.pack(anchor=tk.W)

        # Interference Type Classification legend
        self.interference_class_box = ttk.LabelFrame(int_top, text="Interference Type Classification")
        self.interference_class_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        int_class_inner = ttk.Frame(self.interference_class_box)
        int_class_inner.pack(fill=tk.BOTH, padx=6, pady=6)

        self.interference_legend_table = ttk.Treeview(
            int_class_inner,
            columns=("itype",),
            show="headings",
            height=4,
        )
        self.interference_legend_table.heading("itype", text="Interference Type (Source / Victim)")
        self.interference_legend_table.column("itype", width=190, anchor=tk.CENTER)
        self.interference_legend_table.pack(fill=tk.X)

        # Auto-resize interference legend single column
        def _resize_int_legend(event=None):
            total = max(self.interference_legend_table.winfo_width() - 4, 190)
            self.interference_legend_table.column("itype", width=total)
        self.interference_legend_table.bind("<Configure>", _resize_int_legend)
        _resize_int_legend()

        # Populate interference legend and approximate colors
        irows = [
            ("Inband / Inband", "#FFA600"),
            ("Out of Band / Inband", "#FF6361"),
            ("Inband / Out of Band", "#D359A2"),
            ("Out of Band / Out of Band", "#7D73CA"),
        ]
        for text, color in irows:
            iid = self.interference_legend_table.insert("", tk.END, values=(text,))
            self.interference_legend_table.tag_configure(text, background=color)
            self.interference_legend_table.item(iid, tags=(text,))

        # Results matrix container (holds tksheet or treeview)
        self.interference_matrix = ttk.Frame(self.interference_tab)
        self.interference_matrix.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))

        # Interference buttons
        int_btns = ttk.Frame(self.interference_tab)
        int_btns.pack(fill=tk.X, padx=6, pady=(0, 6))
        self.interference_results_btn = ttk.Button(int_btns, text="Generate Results")
        self.interference_export_btn = ttk.Button(int_btns, text="Export to Excel")
        self.interference_results_btn.pack(side=tk.LEFT)
        self.interference_export_btn.pack(side=tk.LEFT, padx=6)


class App:
    AEDT_VERSION = "2025.2"
    NG_MODE = False  # show AEDT UI when launching if False

    def __init__(self):
        # Launch AEDT desktop
        self.desktop = ansys.aedt.core.launch_desktop(self.AEDT_VERSION, self.NG_MODE, new_desktop=True)

        # Tk window and UI
        self.root = tk.Tk()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self.root)

        # Internal state
        self.emitapp = None
        self.populating_dropdown = False
        self.previous_design = ""
        self.previous_project = ""
        self.global_protection_level = True
        self.protection_levels = {"Global": [30.0, -4.0, -30.0, -104.0]}
        self.tx_radios = []
        self.rx_radios = []
        self.all_colors = []
        self.power_matrix = []
        self._canvas_widgets = {0: None, 1: None}

        # Colors (name -> [ui_color_hex, excel_hex]) matching original mapping
        self.color_dict = {
            "green": ["#7d73ca", "#7d73ca"],
            "yellow": ["#d359a2", "#d359a2"],
            "orange": ["#ff6361", "#ff6361"],
            "red": ["#ffa600", "#ffa600"],
            "white": ["#ffffff", "#ffffff"],
        }

        # Wire events
        self.ui.file_select_btn.configure(command=self.open_file_dialog)
        self.ui.design_name_dropdown.bind("<<ComboboxSelected>>", lambda e: self.design_dropdown_changed())

        self.ui.protection_results_btn.configure(command=self.protection_results)
        self.ui.protection_export_btn.configure(command=self.save_results_excel)

        self.ui.interference_results_btn.configure(command=self.interference_results)
        self.ui.interference_export_btn.configure(command=self.save_results_excel)

        self.ui.radio_specific_levels.configure(command=self.radio_specific)
        self.ui.radio_dropdown.bind("<<ComboboxSelected>>", lambda e: self.radio_dropdown_changed())

        # Legend editing (simple double-click editor on value column)
        self.ui.protection_legend_table.bind("<Double-1>", self._edit_legend_value)

        # Initial button states
        self._set_buttons_enabled(False)

        # Close handler
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        # Prompt for project on startup
        self.open_file_dialog()

    # ---------------------- UI helpers ----------------------
    def _set_warning(self, text):
        self.ui.warnings_var.set(text or "")
        self.ui.warnings.pack()

    def _set_buttons_enabled(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        for b in (
            self.ui.protection_results_btn,
            self.ui.protection_export_btn,
            self.ui.interference_results_btn,
            self.ui.interference_export_btn,
        ):
            try:
                b.configure(state=state)
            except Exception:
                pass

    def _current_tab_index(self):
        return self.ui.tab_widget.index(self.ui.tab_widget.select())

    def _get_legend_values(self):
        # Read second column (value) back from the legend table
        values = []
        for iid in self.ui.protection_legend_table.get_children(""):
            row = self.ui.protection_legend_table.item(iid, option="values")
            try:
                values.append(float(row[1]))
            except Exception:
                values.append(0.0)
        return values

    def _edit_legend_value(self, event):
        # Place an Entry widget over the clicked cell to edit numeric value
        tv = self.ui.protection_legend_table
        region = tv.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = tv.identify_column(event.x)
        if column != "#2":  # only the value column
            return
        row_iid = tv.identify_row(event.y)
        if not row_iid:
            return

        x, y, w, h = tv.bbox(row_iid, column)
        value = tv.set(row_iid, column)
        entry = ttk.Entry(tv)
        entry.insert(0, value)
        entry.place(x=x, y=y, width=w, height=h)

        def commit(event=None):
            new_val = entry.get().strip()
            try:
                float(new_val)
            except ValueError:
                messagebox.showerror("Invalid value", "Please enter a numeric value.")
                entry.focus_set()
                return
            tv.set(row_iid, column, new_val)
            entry.destroy()
            # Persist into protection_levels for current radio or Global
            values = self._get_legend_values()
            index = self.ui.radio_dropdown.get() or "Global"
            self.protection_levels[index] = values

        entry.bind("<Return>", commit)
        entry.bind("<FocusOut>", commit)
        entry.focus_set()

    # ---------------------- File and design selection ----------------------
    def open_file_dialog(self):
        fname = filedialog.askopenfilename(
            title="Select EMIT Project",
            filetypes=[("Ansys Electronics Desktop Files", "*.aedt")],
        )
        if not fname:
            return

        self.ui.file_path_var.set(fname)

        # Close previous project and open specified one
        if self.emitapp is not None:
            try:
                self.emitapp.close_project()
            except Exception:
                pass
            self.emitapp = None

        desktop_proj = self.desktop.load_project(fname)

        # Empty project (no designs)
        if isinstance(desktop_proj, bool):
            self.ui.file_path_var.set("")
            messagebox.showerror("Error: Project missing designs.", "The selected project has no designs. Projects must have at least one EMIT design. See AEDT log for more information.")
            return

        # Locked project
        if desktop_proj.lock_file is None:
            messagebox.showerror("Error: Project already open", "Project is locked. Close or remove the lock before proceeding. See AEDT log for more information.")
            return

        # Populate design dropdown with EMIT designs
        designs = desktop_proj.design_list
        emit_designs = []
        for d in designs:
            design_type = self.desktop.design_type(desktop_proj.project_name, d)
            if design_type == "EMIT":
                emit_designs.append(d)

        self._set_warning(None)
        if not emit_designs:
            self._set_warning("Warning: The project must contain at least one EMIT design.")
            return

        self.populating_dropdown = True
        self.ui.design_name_dropdown.configure(state="readonly", values=emit_designs)
        self.ui.design_name_dropdown.set(emit_designs[0])
        self.populating_dropdown = False

        # Create Emit app on first design
        self.emitapp = ansys.aedt.core.Emit(project=desktop_proj.project_name, design=emit_designs[0])
        self._post_design_selected()

    def _post_design_selected(self):
        # Check radios
        radios = self.emitapp.modeler.components.get_radios()
        self._set_warning(None)
        if len(radios) < 2:
            self._set_warning("Warning: The selected design must contain at least two radios.")

        # Reset radio-specific levels UI state
        self.ui.radio_specific_levels.state(["!disabled"])
        self.ui.radio_dropdown.configure(state="disabled", values=[])
        self.global_protection_level = True
        self.protection_levels = {"Global": self._get_legend_values()}

        # Enable result buttons
        self.ui.protection_results_btn.configure(state="normal")
        self.ui.interference_results_btn.configure(state="normal")
        # Disable exports until we have data
        self.ui.protection_export_btn.configure(state="disabled")
        self.ui.interference_export_btn.configure(state="disabled")

    def design_dropdown_changed(self):
        if self.populating_dropdown:
            return
        design_name = self.ui.design_name_dropdown.get()
        if not design_name:
            return
        self.emitapp = ansys.aedt.core.Emit(project=self.emitapp.project_name, design=design_name)
        self._post_design_selected()
        self.clear_table()

    # ---------------------- Radio-specific legend handling ----------------------
    def radio_specific(self):
        enabled = bool(self.ui.radio_specific_levels.instate(["selected"]))
        self.ui.radio_dropdown.configure(state=("readonly" if enabled else "disabled"))
        self.ui.radio_dropdown.set("")
        if enabled:
            self.emitapp.set_active_design(self.ui.design_name_dropdown.get())
            radios = self.emitapp.modeler.components.get_radios()
            values = self._get_legend_values()
            radios_with_rx = [name for name, obj in radios.items() if obj.has_rx_channels()]
            self.protection_levels = {name: values for name in radios_with_rx}
            self.ui.radio_dropdown.configure(values=radios_with_rx)
            if radios_with_rx:
                self.ui.radio_dropdown.set(radios_with_rx[0])
        else:
            values = self._get_legend_values()
            self.protection_levels = {"Global": values}
        self.global_protection_level = not enabled

    def radio_dropdown_changed(self):
        # Update legend UI with per-radio values
        cur = self.ui.radio_dropdown.get()
        if not cur:
            return
        if cur not in self.protection_levels:
            return
        values = self.protection_levels[cur]
        # Update Treeview value column
        i = 0
        for iid in self.ui.protection_legend_table.get_children(""):
            row_vals = list(self.ui.protection_legend_table.item(iid, "values"))
            row_vals[1] = str(values[i])
            self.ui.protection_legend_table.item(iid, values=tuple(row_vals))
            i += 1

    # ---------------------- Save Excel export ----------------------
    def save_results_excel(self):
        default_name = "Protection Level Classification" if self._current_tab_index() == 0 else "Interference Type Classification"
        table = self.ui.protection_matrix if self._current_tab_index() == 0 else self.ui.interference_matrix
        fname = filedialog.asksaveasfilename(title="Save Scenario Matrix", defaultextension=".xlsx", initialfile=default_name, filetypes=[("xlsx", "*.xlsx")])
        
        if not fname:
            return

        workbook = openpyxl.Workbook()
        ws = workbook.active
        header = self.tx_radios[:]
        header.insert(0, "Tx/Rx")
        ws.append(header)

        # all_colors indexed [col][row], power_matrix similarly
        for r in range(len(self.rx_radios)):
            ws.cell(row=r + 2, column=1, value=str(self.rx_radios[r]))
            for c in range(len(self.tx_radios)):
                text = str(self.power_matrix[c][r])
                ws.cell(row=r + 2, column=c + 2, value=text)
                color_name = self.all_colors[c][r]
                hex_color = self.color_dict.get(color_name, ["#ffffff", "#ffffff"])[1].lstrip("#")
                ws.cell(row=r + 2, column=c + 2).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        workbook.save(fname)

    # ---------------------- Compute results ----------------------
    def interference_results(self):
        checks = [
            bool(self.ui.in_in_var.get()),
            bool(self.ui.out_in_var.get()),
            bool(self.ui.in_out_var.get()),
            bool(self.ui.out_out_var.get()),
        ]
        filters = [
            "TxFundamental:In-band",
            ["TxHarmonic/Spurious:In-band", "Intermod:In-band", "Broadband:In-band"],
            "TxFundamental:Out-of-band",
            ["TxHarmonic/Spurious:Out-of-band", "Intermod:Out-of-band", "Broadband:Out-of-band"],
        ]
        flt = [f for (f, v) in zip(filters, checks) if v]

        if not self.ui.file_path_var.get() or not self.ui.design_name_dropdown.get():
            return

        if self.previous_design != self.ui.design_name_dropdown.get() or self.previous_project != self.ui.file_path_var.get():
            self.previous_design = self.ui.design_name_dropdown.get()
            self.previous_project = self.ui.file_path_var.get()
            self.emitapp.set_active_design(self.ui.design_name_dropdown.get())

            if self.emitapp.read_only_variable():
                messagebox.showerror("Writing Error", "An error occurred while writing to the file. Is it readonly? Disk full? See AEDT log for more information.")
                return

            rev = self.emitapp.results.analyze()
            self.tx_interferer = InterfererType().TRANSMITTERS
            self.rx_radios = rev.get_receiver_names()
            self.tx_radios = rev.get_interferer_names(self.tx_interferer)
            if self.tx_radios is None or self.rx_radios is None:
                return
            self.rev = rev

        domain = self.emitapp.results.interaction_domain()
        self.all_colors, self.power_matrix = self.rev.interference_type_classification(domain, interferer_type=self.tx_interferer, use_filter=True, filter_list=flt)
        try:
            self.emitapp.save_project()
        except Exception:
            pass
        self.populate_table()

    def protection_results(self):
        checks = [
            bool(self.ui.damage_var.get()),
            bool(self.ui.overload_var.get()),
            bool(self.ui.intermodulation_var.get()),
            bool(self.ui.desensitization_var.get()),
        ]
        filters = ["damage", "overload", "intermodulation", "desensitization"]
        flt = [f for (f, v) in zip(filters, checks) if v]

        if not self.ui.file_path_var.get() or not self.ui.design_name_dropdown.get():
            return

        if self.previous_design != self.ui.design_name_dropdown.get() or self.previous_project != self.ui.file_path_var.get():
            self.previous_design = self.ui.design_name_dropdown.get()
            self.previous_project = self.ui.file_path_var.get()
            self.emitapp.set_active_design(self.ui.design_name_dropdown.get())

            if self.emitapp.read_only_variable():
                messagebox.showerror("Writing Error", "An error occurred while writing to the file. Is it readonly? Disk full? See AEDT log for more information.")
                return

            self.tx_interferer = InterfererType().TRANSMITTERS
            rev = self.emitapp.results.analyze()
            self.rx_radios = rev.get_receiver_names()
            self.tx_radios = rev.get_interferer_names(self.tx_interferer)
            self.rev = rev
            if self.tx_radios is None or self.rx_radios is None:
                return

        domain = self.emitapp.results.interaction_domain()
        global_levels = self.protection_levels.get("Global", self._get_legend_values())
        self.all_colors, self.power_matrix = self.rev.protection_level_classification(
            domain=domain,
            interferer_type=self.tx_interferer,
            global_protection_level=self.global_protection_level,
            global_levels=global_levels,
            protection_levels=self.protection_levels,
            use_filter=True,
            filter_list=flt,
        )
        self.populate_table()

    # ---------------------- Table rendering ----------------------
    def populate_table(self):
        idx = self._current_tab_index()
        if idx == 0:
            canvas = self.ui.protection_matrix
            export_btn = self.ui.protection_export_btn
        else:
            canvas = self.ui.interference_matrix
            export_btn = self.ui.interference_export_btn

        # Clear any previously rendered child widgets in the container
        for child in canvas.winfo_children():
            try:
                child.destroy()
            except Exception:
                pass

        # Ensure we have data to display
        num_cols = len(self.tx_radios)
        num_rows = len(self.rx_radios)
        if num_cols == 0 or num_rows == 0:
            export_btn.configure(state="disabled")
            return

        # Draw a resizable grid on Canvas with per-cell backgrounds
        parent = canvas
        cnv = tk.Canvas(parent, highlightthickness=0, background="white")
        cnv.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))
        self._canvas_widgets[idx] = cnv

        def draw_table(event=None):
            cnv.delete("all")
            W = max(cnv.winfo_width(), 200)
            H = max(cnv.winfo_height(), 150)
            margin_left = 120
            margin_top = 26
            grid_x0 = 1
            grid_y0 = 1
            grid_x1 = W - 2
            grid_y1 = H - 2
            # headers area
            header_w = margin_left
            header_h = margin_top
            # cell area
            cell_x0 = grid_x0 + header_w
            cell_y0 = grid_y0 + header_h
            cell_w = max(grid_x1 - cell_x0, 50)
            cell_h = max(grid_y1 - cell_y0, 50)

            # Compute per-column/row sizes
            col_w = cell_w / max(num_cols, 1)
            row_h = cell_h / max(num_rows, 1)

            # Draw Rx headers (rows)
            for r in range(num_rows):
                y0 = cell_y0 + r * row_h
                y1 = y0 + row_h
                cnv.create_rectangle(grid_x0, y0, grid_x0 + header_w, y1, fill="#f2f2f2", outline="#cccccc")
                cnv.create_text(grid_x0 + 6, (y0 + y1) / 2, text=str(self.rx_radios[r]), anchor="w")

            # Draw Tx headers (columns)
            for c in range(num_cols):
                x0 = cell_x0 + c * col_w
                x1 = x0 + col_w
                cnv.create_rectangle(x0, grid_y0, x1, grid_y0 + header_h, fill="#f2f2f2", outline="#cccccc")
                cnv.create_text((x0 + x1) / 2, grid_y0 + header_h / 2, text=str(self.tx_radios[c]), anchor="center")

            # Draw cells with per-cell background and text
            for r in range(num_rows):
                for c in range(num_cols):
                    x0 = cell_x0 + c * col_w
                    y0 = cell_y0 + r * row_h
                    x1 = x0 + col_w
                    y1 = y0 + row_h
                    color_name = self.all_colors[c][r]
                    fill_hex = self.color_dict.get(color_name, ["#ffffff", "#ffffff"])[0]
                    cnv.create_rectangle(x0, y0, x1, y1, fill=fill_hex, outline="#ffffff")
                    txt = str(self.power_matrix[c][r])
                    cnv.create_text((x0 + x1) / 2, (y0 + y1) / 2, text=txt, anchor="center")

            cnv.update_idletasks()

        cnv.bind("<Configure>", draw_table)
        draw_table()
        self.root.update_idletasks()
        cnv.update_idletasks()
        self.image_capture = ImageGrab.grab(bbox=(cnv.winfo_rootx(), cnv.winfo_rooty(), cnv.winfo_rootx() + cnv.winfo_width(), cnv.winfo_rooty() + cnv.winfo_height()))

        export_btn.configure(state="normal")

    def clear_table(self):
        if self._current_tab_index() == 0:
            canvas = self.ui.protection_matrix
            export_btn = self.ui.protection_export_btn
        else:
            canvas = self.ui.interference_matrix
            export_btn = self.ui.interference_export_btn

        # Clear any table widgets in the container for this tab
        for child in canvas.winfo_children():
            try:
                child.destroy()
            except Exception:
                pass
        idx = self._current_tab_index()
        self._canvas_widgets[idx] = None
        export_btn.configure(state="disabled")

    # ---------------------- Close ----------------------
    def on_close(self):
        messagebox.showinfo("Closing GUI", "Closing AEDT. Wait for the GUI to close on its own.")
        try:
            if self.emitapp:
                self.emitapp.close_project()
                self.emitapp.close_desktop()
            else:
                self.desktop.release_desktop(True, True)
        except Exception:
            pass
        self.root.destroy()

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    App().run()
